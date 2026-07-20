"""Student management service — CRUD, bulk import (Excel), export, audit."""
from __future__ import annotations

import io
import re
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

from app.core.logging import get_logger
from app.core.security import hash_password
from app.repositories import admin_activity as audit
from app.repositories import students as students_repo
from app.repositories import users as users_repo
from app.services.auth_service import ROLE_STUDENT

log = get_logger(__name__)

# Audit action constants (student-scoped)
ACTION_STUDENT_CREATED = "STUDENT_CREATED"
ACTION_STUDENT_UPDATED = "STUDENT_UPDATED"
ACTION_STUDENT_DELETED = "STUDENT_DELETED"
ACTION_STUDENT_STATUS = "STUDENT_STATUS_CHANGED"
ACTION_STUDENT_PASSWORD_RESET = "STUDENT_PASSWORD_RESET"
ACTION_STUDENT_BULK_IMPORT = "STUDENT_BULK_IMPORT"
ACTION_STUDENT_EXPORT = "STUDENT_EXPORT"

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

EXCEL_COLUMNS = [
    "Roll Number",
    "Full Name",
    "Email",
    "Phone",
    "Department",
    "Branch",
    "Year",
    "Gender",
    "Status",
]

_COLUMN_ALIASES = {
    "roll number": "rollNumber",
    "rollno": "rollNumber",
    "roll no": "rollNumber",
    "roll": "rollNumber",
    "full name": "fullName",
    "name": "fullName",
    "email": "email",
    "phone": "phone",
    "mobile": "phone",
    "department": "department",
    "dept": "department",
    "branch": "branch",
    "year": "year",
    "gender": "gender",
    "status": "status",
}


class StudentServiceError(Exception):
    def __init__(self, message: str, status_code: int = 400):
        super().__init__(message)
        self.status_code = status_code


def _ensure_indexes() -> None:
    try:
        students_repo.ensure_indexes()
        audit.ensure_indexes()
    except Exception as e:  # noqa: BLE001
        log.warning("Student index bootstrap skipped: %s", e)


def _log(action: str, actor: dict[str, Any], *, target: dict[str, Any] | None = None,
         actor_ip: str | None = None, metadata: dict[str, Any] | None = None) -> None:
    audit.log_action(
        action=action,
        actor_user_id=actor.get("userId"),
        actor_email=actor.get("email"),
        target_user_id=(target or {}).get("studentId"),
        target_email=(target or {}).get("email"),
        ip=actor_ip,
        metadata=metadata,
    )


def _validate_new_student(payload: dict[str, Any]) -> None:
    if not (payload.get("fullName") or "").strip():
        raise StudentServiceError("Full name is required.", 422)
    if not (payload.get("rollNumber") or "").strip():
        raise StudentServiceError("Roll number is required.", 422)
    email = (payload.get("email") or "").strip()
    if not email or not EMAIL_RE.match(email):
        raise StudentServiceError("A valid email is required.", 422)


# ─── Listing ────────────────────────────────────────────────────────────
def list_students(**kwargs) -> dict[str, Any]:
    _ensure_indexes()
    rows, total = students_repo.list_students(**kwargs)
    fac = students_repo.facets()
    return {"items": rows, "total": total, "facets": fac}


def get_student(student_id: str) -> dict[str, Any]:
    s = students_repo.get_by_id(student_id)
    if not s:
        raise StudentServiceError("Student not found", 404)
    return s


# ─── Create ─────────────────────────────────────────────────────────────
def create_student(payload: dict[str, Any], *, actor: dict[str, Any],
                   actor_ip: str | None = None) -> dict[str, Any]:
    _ensure_indexes()
    _validate_new_student(payload)
    email = payload["email"].strip().lower()
    roll = payload["rollNumber"].strip().upper()

    if students_repo.get_by_email(email) or users_repo.get_by_email(email):
        raise StudentServiceError("A user with this email already exists.", 409)
    if students_repo.get_by_roll(roll):
        raise StudentServiceError("A student with this roll number already exists.", 409)

    student = students_repo.create_student(
        roll_number=roll,
        full_name=payload["fullName"],
        email=email,
        phone=payload.get("phone", ""),
        department=payload.get("department", ""),
        branch=payload.get("branch", ""),
        year=payload.get("year", ""),
        gender=payload.get("gender", ""),
        status=payload.get("status") or students_repo.STATUS_ACTIVE,
        password_hash=hash_password(roll),  # default password = roll number
        must_change_password=True,
        created_by=actor.get("userId"),
    )
    _log(ACTION_STUDENT_CREATED, actor, target=student, actor_ip=actor_ip,
         metadata={"rollNumber": roll})
    return student


# ─── Update ─────────────────────────────────────────────────────────────
def update_student(student_id: str, patch: dict[str, Any], *,
                   actor: dict[str, Any], actor_ip: str | None = None) -> dict[str, Any]:
    target = get_student(student_id)
    # Guard duplicates on email/roll changes
    new_email = (patch.get("email") or "").strip().lower() if patch.get("email") else None
    if new_email and new_email != target["email"]:
        if students_repo.get_by_email(new_email) or users_repo.get_by_email(new_email):
            raise StudentServiceError("Another user already uses this email.", 409)
    new_roll = (patch.get("rollNumber") or "").strip().upper() if patch.get("rollNumber") else None
    if new_roll and new_roll != target["rollNumber"]:
        if students_repo.get_by_roll(new_roll):
            raise StudentServiceError("Another student already uses this roll number.", 409)

    updated = students_repo.update_student(student_id, patch) or target
    _log(ACTION_STUDENT_UPDATED, actor, target=updated, actor_ip=actor_ip,
         metadata={k: v for k, v in patch.items() if k != "passwordHash"})
    return updated


def set_status(student_id: str, status: str, *, actor: dict[str, Any],
               actor_ip: str | None = None) -> dict[str, Any]:
    target = get_student(student_id)
    try:
        updated = students_repo.set_status(student_id, status)
    except ValueError as e:
        raise StudentServiceError(str(e), 422)
    _log(ACTION_STUDENT_STATUS, actor, target=updated or target, actor_ip=actor_ip,
         metadata={"status": status, "previous": target.get("status")})
    return updated  # type: ignore[return-value]


def soft_delete(student_id: str, *, actor: dict[str, Any],
                actor_ip: str | None = None) -> dict[str, Any]:
    return set_status(student_id, students_repo.STATUS_DELETED, actor=actor, actor_ip=actor_ip)


def reset_password(student_id: str, *, actor: dict[str, Any],
                   actor_ip: str | None = None) -> dict[str, Any]:
    target = get_student(student_id)
    roll = target["rollNumber"]
    students_repo.set_password(student_id, hash_password(roll), must_change=True)
    _log(ACTION_STUDENT_PASSWORD_RESET, actor, target=target, actor_ip=actor_ip,
         metadata={"defaultTo": "rollNumber"})
    return {"studentId": student_id, "defaultPassword": roll}


# ─── Bulk import ────────────────────────────────────────────────────────
def _normalize_headers(headers: list[str]) -> list[str | None]:
    out: list[str | None] = []
    for h in headers:
        key = (str(h or "")).strip().lower()
        out.append(_COLUMN_ALIASES.get(key))
    return out


def parse_excel(data: bytes) -> dict[str, Any]:
    """Parse an uploaded xlsx/xls and return preview rows + issues (no DB writes)."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise StudentServiceError(f"Unable to read Excel file: {e}", 400)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise StudentServiceError("The Excel file is empty.", 422)
    headers_raw = [str(h).strip() if h is not None else "" for h in rows[0]]
    mapping = _normalize_headers(headers_raw)
    if "rollNumber" not in mapping or "email" not in mapping or "fullName" not in mapping:
        raise StudentServiceError(
            "Excel must contain at least: Roll Number, Full Name, Email.", 422
        )

    valid: list[dict[str, Any]] = []
    duplicates_in_file: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    seen_emails: set[str] = set()
    seen_rolls: set[str] = set()

    for idx, raw in enumerate(rows[1:], start=2):
        if not raw or all(c in (None, "") for c in raw):
            continue
        entry: dict[str, Any] = {}
        for i, key in enumerate(mapping):
            if key is None or i >= len(raw):
                continue
            val = raw[i]
            entry[key] = "" if val is None else str(val).strip()

        entry.setdefault("status", students_repo.STATUS_ACTIVE)
        errors: list[str] = []
        email = (entry.get("email") or "").lower()
        roll = (entry.get("rollNumber") or "").upper()
        entry["email"] = email
        entry["rollNumber"] = roll

        if not entry.get("fullName"):
            errors.append("Full name is required")
        if not roll:
            errors.append("Roll number is required")
        if not email:
            errors.append("Email is required")
        elif not EMAIL_RE.match(email):
            errors.append("Invalid email format")

        if errors:
            invalid.append({"row": idx, "data": entry, "errors": errors})
            continue

        if email in seen_emails or roll in seen_rolls:
            duplicates_in_file.append({"row": idx, "data": entry, "reason": "Duplicate in file"})
            continue
        seen_emails.add(email)
        seen_rolls.add(roll)
        entry["row"] = idx
        valid.append(entry)

    # Check duplicates against DB
    existing: list[dict[str, Any]] = []
    to_import: list[dict[str, Any]] = []
    for e in valid:
        if students_repo.get_by_email(e["email"]) or users_repo.get_by_email(e["email"]):
            existing.append({"row": e["row"], "data": e, "reason": "Email already exists"})
        elif students_repo.get_by_roll(e["rollNumber"]):
            existing.append({"row": e["row"], "data": e, "reason": "Roll number already exists"})
        else:
            to_import.append(e)

    return {
        "columns": EXCEL_COLUMNS,
        "totalRows": len(rows) - 1,
        "toImport": to_import,
        "duplicatesInFile": duplicates_in_file,
        "existingInDb": existing,
        "invalidRows": invalid,
    }


def bulk_import(rows: list[dict[str, Any]], *, actor: dict[str, Any],
                actor_ip: str | None = None) -> dict[str, Any]:
    _ensure_indexes()
    imported: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    for entry in rows:
        try:
            _validate_new_student(entry)
            email = entry["email"].strip().lower()
            roll = entry["rollNumber"].strip().upper()
            if (
                students_repo.get_by_email(email)
                or students_repo.get_by_roll(roll)
                or users_repo.get_by_email(email)
            ):
                skipped.append({"data": entry, "reason": "Duplicate email or roll number"})
                continue
            student = students_repo.create_student(
                roll_number=roll,
                full_name=entry["fullName"],
                email=email,
                phone=entry.get("phone", ""),
                department=entry.get("department", ""),
                branch=entry.get("branch", ""),
                year=entry.get("year", ""),
                gender=entry.get("gender", ""),
                status=entry.get("status") or students_repo.STATUS_ACTIVE,
                password_hash=hash_password(roll),
                must_change_password=True,
                created_by=actor.get("userId"),
            )
            imported.append(student)
        except StudentServiceError as e:
            skipped.append({"data": entry, "reason": str(e)})
        except Exception as e:  # noqa: BLE001
            log.exception("Bulk import row failed")
            skipped.append({"data": entry, "reason": f"Unexpected error: {e}"})

    _log(ACTION_STUDENT_BULK_IMPORT, actor, actor_ip=actor_ip,
         metadata={"imported": len(imported), "skipped": len(skipped)})
    return {"imported": imported, "skipped": skipped,
            "counts": {"imported": len(imported), "skipped": len(skipped)}}


# ─── Export / template ──────────────────────────────────────────────────
def _write_workbook(rows: Iterable[dict[str, Any]], sheet_name: str = "Students") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(EXCEL_COLUMNS)
    for r in rows:
        ws.append([
            r.get("rollNumber", ""),
            r.get("fullName", ""),
            r.get("email", ""),
            r.get("phone", ""),
            r.get("department", ""),
            r.get("branch", ""),
            r.get("year", ""),
            r.get("gender", ""),
            r.get("status", ""),
        ])
    for i, _ in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[chr(64 + i)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_students(*, actor: dict[str, Any], actor_ip: str | None = None,
                    ids: list[str] | None = None, filters: dict[str, Any] | None = None) -> bytes:
    filters = filters or {}
    rows, _ = students_repo.list_students(
        search=filters.get("search", ""),
        department=filters.get("department", ""),
        branch=filters.get("branch", ""),
        year=filters.get("year", ""),
        status=filters.get("status", ""),
        include_deleted=True,
        skip=0,
        limit=10000,
        ids=ids,
    )
    _log(ACTION_STUDENT_EXPORT, actor, actor_ip=actor_ip,
         metadata={"count": len(rows), "byIds": bool(ids)})
    return _write_workbook(rows)


def import_template() -> bytes:
    example = [{
        "rollNumber": "22CSE001",
        "fullName": "Surendra",
        "email": "surendra@example.edu",
        "phone": "9876543210",
        "department": "Engineering",
        "branch": "CSE",
        "year": "2",
        "gender": "Male",
        "status": "ACTIVE",
    }]
    return _write_workbook(example, sheet_name="Template")
