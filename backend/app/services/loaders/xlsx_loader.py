from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import LoadedChunk, clean_text


def _row_to_text(headers: list[str], row: tuple) -> str:
    parts: list[str] = []
    for header, value in zip(headers, row):
        if value is None or str(value).strip() == "":
            continue
        if header:
            parts.append(f"{header}: {value}")
        else:
            parts.append(str(value))
    return clean_text(" | ".join(parts))


def load_xlsx(path: Path) -> list[LoadedChunk]:
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except (InvalidFileException, KeyError, OSError) as e:
        raise ValueError(f"Corrupted or invalid XLSX: {path.name}") from e

    chunks: list[LoadedChunk] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        # Skip fully-empty sheets
        if not any(any(c is not None and str(c).strip() != "" for c in r) for r in rows):
            continue

        # First non-empty row = headers.
        headers: list[str] = []
        data_start = 0
        for i, r in enumerate(rows):
            if any(c is not None and str(c).strip() != "" for c in r):
                headers = [clean_text(str(c)) if c is not None else "" for c in r]
                data_start = i + 1
                break

        row_no = 0
        for r in rows[data_start:]:
            if not any(c is not None and str(c).strip() != "" for c in r):
                continue
            row_no += 1
            text = _row_to_text(headers, r)
            if not text:
                continue
            chunks.append(
                LoadedChunk(
                    text=text,
                    metadata={
                        "document_name": path.name,
                        "sheet_name": sheet.title,
                        "row_number": row_no,
                        "paragraph_number": row_no,
                        "source_type": "xlsx",
                    },
                )
            )
    wb.close()
    return chunks
